import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import openpyxl
from openpyxl.styles import PatternFill
import random
import os
from datetime import datetime
import base64

st.set_page_config(layout="wide")

st.title("ЯП-комбайн")

def get_cell_color(workbook, sheet_name, row, col):
    try:
        worksheet = workbook[sheet_name]
        cell = worksheet.cell(row=row, column=col)
        fill = cell.fill
        if hasattr(fill, 'fill'):
            actual_fill = fill.fill
        else:
            actual_fill = fill
        if hasattr(actual_fill, 'fgColor') and hasattr(actual_fill.fgColor, 'rgb'):
            rgb = actual_fill.fgColor.rgb
            if rgb:
                rgb_lower = rgb.lower()
                if rgb_lower in ['ffff0000', '00ff0000', 'ff0000', 'ff0000ff']:
                    return 'red'
                elif rgb_lower in ['ffffff00', '00ffff00', 'ffff00']:
                    return 'yellow'
        return None
    except Exception:
        return None

def get_download_link(file_path, file_name):
    with open(file_path, "rb") as f:
        bytes_data = f.read()
    b64 = base64.b64encode(bytes_data).decode()
    href = f'<a href="data:application/octet-stream;base64,{b64}" download="{file_name}">Скачать {file_name}</a>'
    return href

def create_html_table(df, workbook, sheet_name):
    html = """
    <style>
        .table-container::-webkit-scrollbar {
            height: 12px;
        }
        .table-container::-webkit-scrollbar-thumb {
            background-color: #888;
            border-radius: 6px;
        }
        .table-container::-webkit-scrollbar-thumb:hover {
            background-color: #555;
        }
    </style>
    <div class="table-container" style="overflow-x: auto; width: 100%;" id="tableContainer">
        <table style="border-collapse: collapse; width: 100%; min-width: max-content;">
    """
    html += "<tr style='background-color: #f2f2f2;'>"
    html += "<th style='border: 1px solid #ddd; padding: 8px; position: sticky; left: 0; background-color: #f2f2f2; z-index: 1;'></th>"
    for col in df.columns:
        html += f"<th style='border: 1px solid #ddd; padding: 8px;'>{col}</th>"
    html += "</tr>"
    for i, (index, row) in enumerate(df.iterrows()):
        html += "<tr>"
        html += f"<td style='border: 1px solid #ddd; padding: 8px; font-weight: bold; position: sticky; left: 0; background-color: #ffffff; z-index: 1;'>{index}</td>"
        for j, value in enumerate(row):
            color = get_cell_color(workbook, sheet_name, i + 2, j + 2)
            style = "border: 1px solid #ddd; padding: 8px;"
            if color == 'red':
                style += "background-color: #ffcccc;"
            elif color == 'yellow':
                style += "background-color: #ffffcc;"
            html += f"<td style='{style}'>{value}</td>"
        html += "</tr>"
    html += "</table>"
    html += "</div>"
    html += """
    <script>
        document.addEventListener("DOMContentLoaded", function() {
            var container = document.getElementById("tableContainer");
            container.scrollLeft = container.scrollWidth;
        });
    </script>
    """
    return html

def create_top5_table(df):
    # Исключаем последнюю строку (суммарное количество)
    df_without_total = df.iloc[:-1]
    
    last_column = df_without_total.columns[-1]
    temp_df = pd.DataFrame({
        "Название": df_without_total.index,
        "Кол-во голосов": df_without_total[last_column]
    })
    top5_df = temp_df.sort_values(by="Кол-во голосов", ascending=False).head(5)
    top5_df["Место"] = range(1, len(top5_df) + 1)
    top5_df = top5_df[["Место", "Название", "Кол-во голосов"]]
    
    html = '<table style="width: 100%; max-width: 500px; border-collapse: collapse; margin-bottom: 20px;">'
    html += '<thead><tr style="background-color: #f2f2f2;">'
    html += '<th style="border: 1px solid #ddd; padding: 8px; text-align: left;">Место</th>'
    html += '<th style="border: 1px solid #ddd; padding: 8px; text-align: left;">Название</th>'
    html += '<th style="border: 1px solid #ddd; padding: 8px; text-align: left;">Кол-во голосов</th>'
    html += '</tr></thead><tbody>'
    
    for _, row in top5_df.iterrows():
        html += '<tr>'
        html += f'<td style="border: 1px solid #ddd; padding: 8px;">{row["Место"]}</td>'
        html += f'<td style="border: 1px solid #ddd; padding: 8px;">{row["Название"]}</td>'
        html += f'<td style="border: 1px solid #ddd; padding: 8px;">{row["Кол-во голосов"]}</td>'
        html += '</tr>'
    html += '</tbody></table>'
    
    return html
    
# Загрузка файла
default_file = "output_highlighted.xlsx"
uploaded_file = None
if os.path.exists(default_file):
    st.markdown(f"Файл по умолчанию доступен: {get_download_link(default_file, default_file)}", unsafe_allow_html=True)
else:
    uploaded_file = st.file_uploader("Загрузите Excel файл", type=["xlsx", "xls"])

if uploaded_file is not None or os.path.exists(default_file):
    try:
        if uploaded_file is None and os.path.exists(default_file):
            df = pd.read_excel(default_file)
            wb = openpyxl.load_workbook(default_file)
        else:
            df = pd.read_excel(uploaded_file)
            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(uploaded_file.read()))
        
        df.set_index(df.columns[0], inplace=True)
        sheet_name = wb.sheetnames[0]

        at_date = wb.sheetnames[-1]            # !!!
        
        # Таблица Топ-5
        st.subheader(f"Топ-5! ({sheet_name})")    # !!!
        
        top5_html = create_top5_table(df)
        st.markdown(top5_html, unsafe_allow_html=True)
        
        # Превью Excel-файла
        st.subheader("Превью Excel-файла")
        html_table = create_html_table(df, wb, sheet_name)
        st.markdown(html_table, unsafe_allow_html=True)
        
        # Добавляем графический файл voting_heatmap.png
        st.subheader("Коррелятор")
        st.markdown(f"За какие пары рассказов голосуют одновременно чаще всего")
        st.markdown("<div style='margin-top: 20px;'></div>", unsafe_allow_html=True)
        
        image_path = "voting_heatmap.png"
        if os.path.exists(image_path):
            with open(image_path, "rb") as image_file:
                encoded_image = base64.b64encode(image_file.read()).decode()
            image_html = f"""
            <style>
                .image-container {{
                    width: 80%;
                    text-align: center;
                }}
                .image-container img {{
                    max-width: 80%;
                    height: auto;
                    cursor: pointer;
                }}
                .fullscreen {{
                    display: none;
                    position: fixed;
                    top: 0;
                    left: 0;
                    width: 100%;
                    height: 100%;
                    background: rgba(0, 0, 0, 0.9);
                    z-index: 9999;
                    justify-content: center;
                    align-items: center;
                }}
                .fullscreen img {{
                    max-width: 90%;
                    max-height: 90%;
                }}
            </style>
            <div class="image-container">
                <img src="data:image/png;base64,{encoded_image}" onclick="openFullscreen(this)">
            </div>
            <div class="fullscreen" id="fullscreen">
                <img src="data:image/png;base64,{encoded_image}" onclick="closeFullscreen()">
            </div>
            <script>
                function openFullscreen(element) {{
                    document.getElementById("fullscreen").style.display = "flex";
                }}
                function closeFullscreen() {{
                    document.getElementById("fullscreen").style.display = "none";
                }}
            </script>
            """
            st.markdown(image_html, unsafe_allow_html=True)
        else:
            st.warning("Файл voting_heatmap.png не найден в директории скрипта.")

        st.subheader("Графики")
        st.markdown(f"Красная точка - один голос")
        # Выбор характеристик и типа графика
        params = st.multiselect("Выберите рассказы", df.index.tolist())
        chart_type = st.selectbox("Выберите тип графика", ["Линейный", "Столбчатый", "Точечный", "Площадной"])
        
        if params:
            fig = go.Figure()
            line_colors = ['#0000FF', '#FF0000', '#00FF00', '#00FFFF', '#FF00FF', '#FFFF00', '#000000']
            if len(params) > len(line_colors):
                random_colors = [f'#{random.randint(0, 255):02x}{random.randint(0, 255):02x}{random.randint(0, 255):02x}' 
                               for _ in range(len(params) - len(line_colors))]
                line_colors.extend(random_colors)
            
            for i, param in enumerate(params):
                color = line_colors[i % len(line_colors)]
                x_data = df.columns
                y_data = df.loc[param]
                
                if chart_type == "Линейный":
                    fig.add_trace(go.Scatter(x=x_data, y=y_data, mode='lines', name=param, line=dict(color=color, width=2)))
                    point_colors = [get_cell_color(wb, sheet_name, df.index.get_loc(param) + 2, col) for col in range(2, len(df.columns) + 2)]
                    red_x = [x for x, pc in zip(x_data, point_colors) if pc == 'red']
                    red_y = [y for y, pc in zip(y_data, point_colors) if pc == 'red']
                    if red_x:
                        fig.add_trace(go.Scatter(x=red_x, y=red_y, mode='markers', name=f'{param} (red points)', 
                                               marker=dict(color='red', size=10, line=dict(color='black', width=1)), showlegend=False))
                
                elif chart_type == "Столбчатый":
                    fig.add_trace(go.Bar(x=x_data, y=y_data, name=param, marker_color=color, width=0.8))
                
                elif chart_type == "Точечный":
                    fig.add_trace(go.Scatter(x=x_data, y=y_data, mode='markers', name=param, 
                                           marker=dict(color=color, size=8, line=dict(color='black', width=1))))
                    point_colors = [get_cell_color(wb, sheet_name, df.index.get_loc(param) + 2, col) for col in range(2, len(df.columns) + 2)]
                    red_x = [x for x, pc in zip(x_data, point_colors) if pc == 'red']
                    red_y = [y for y, pc in zip(y_data, point_colors) if pc == 'red']
                    if red_x:
                        fig.add_trace(go.Scatter(x=red_x, y=red_y, mode='markers', name=f'{param} (red points)', 
                                               marker=dict(color='red', size=10, line=dict(color='black', width=1)), showlegend=False))
                
                elif chart_type == "Площадной":
                    r = int(color.lstrip('#')[0:2], 16)
                    g = int(color.lstrip('#')[2:4], 16)
                    b = int(color.lstrip('#')[4:6], 16)
                    fillcolor = f'rgba({r}, {g}, {b}, 0.3)'
                    fig.add_trace(go.Scatter(x=x_data, y=y_data, mode='lines', name=param, fill='tozeroy', 
                                           line=dict(color=color, width=2), fillcolor=fillcolor))
            
            fig.update_layout(
                xaxis_title="Время",
                yaxis_title="Значение",
                title="Графики выбранных рассказов",
                xaxis=dict(tickangle=45, tickmode='auto', nticks=10 if len(df.columns) > 10 else None),
                legend=dict(yanchor="bottom", y=-0.4, xanchor="center", x=0.5, orientation="h", font=dict(size=10)),
                height=600,
                margin=dict(l=50, r=50, t=50, b=100),
                showlegend=True
            )
            fig.update_layout(
                xaxis=dict(showgrid=True, gridcolor='rgba(200, 200, 200, 0.7)', griddash='dash'),
                yaxis=dict(showgrid=True, gridcolor='rgba(200, 200, 0.7)', griddash='dash')
            )
            try:
                dates = [datetime.strptime(str(x), '%d.%m %H:%M') for x in x_data]
                days = [d.date() for d in dates]
                unique_days = sorted(set(days))
                shapes = []
                for i, day in enumerate(unique_days):
                    day_indices = [j for j, d in enumerate(days) if d == day]
                    if day_indices:
                        start_idx = day_indices[0]
                        end_idx = day_indices[-1] + 1 if day_indices[-1] < len(x_data) - 1 else day_indices[-1]
                        fill_color = '#FFFFE0' if i % 2 == 0 else 'white'
                        shapes.append(dict(
                            type="rect", x0=start_idx, x1=end_idx, y0=0, y1=1, yref="paper",
                            fillcolor=fill_color, opacity=0.9, layer="below", line_width=0
                        ))
                fig.update_layout(shapes=shapes)
            except ValueError:
                shapes = []
                for i in range(len(x_data)):
                    fill_color = '#FFFFE0' if i % 2 == 0 else 'white'
                    shapes.append(dict(
                        type="rect", x0=i, x1=i + 1, y0=0, y1=1, yref="paper",
                        fillcolor=fill_color, opacity=0.3, layer="below", line_width=0
                    ))
                fig.update_layout(shapes=shapes)
            
            st.plotly_chart(fig, use_container_width=True)
        
        else:
            st.write("Пожалуйста, выберите хотя бы одну характеристику.")
    except Exception as e:
        st.error(f"Ошибка: {str(e)}")
